import sys
import openpyxl

PATH = "/Users/nagavenkatasatyas/dev/personal_projects/uprep/Chapter List JEE and NEET Class XI & XII.xlsx"

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print("SHEETS:", wb.sheetnames)
for ws in wb.worksheets:
    print("\n===== SHEET:", repr(ws.title), "dims:", ws.max_row, "x", ws.max_column, "=====")
    n = 0
    for row in ws.iter_rows(values_only=True):
        vals = [("" if v is None else str(v)).strip() for v in row]
        if any(vals):
            print(n, "|", " | ".join(vals[:10]))
        n += 1
        if n >= 25:
            break
