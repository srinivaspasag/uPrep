#!/usr/bin/env python3
"""
Import the "Chapter List JEE and NEET" workbook into UPrep as course structure
+ videos.

Mapping:
  Sheet (e.g. "Physics XI")  -> top-level CMDS folder
    Chapter (col A)          -> sub-folder
      Session (col B)        -> sub-sub-folder (e.g. "SESSION 1")
        Topic row (col C+D)  -> a VIDEO added by URL (Vimeo/YouTube), filed in
                                the session folder (linkType ADDED embed).
  (Rows with no session value are filed directly in the chapter folder.)

Idempotent: existing folders (by name under parent) and existing videos (by
title within a chapter folder) are reused/skipped, so re-running or scaling up
does not create duplicates.

Usage:
  UPREP_USER=dummystaff UPREP_PASS=... python3 scripts/import_chapters.py summary
  UPREP_USER=dummystaff UPREP_PASS=... python3 scripts/import_chapters.py import \
      --sheet "Physics XI" --max-chapters 2
"""
import argparse
import os
import sys
import time
import openpyxl
import requests

XLSX = "/Users/nagavenkatasatyas/dev/personal_projects/uprep/Chapter List JEE and NEET Class XI & XII.xlsx"
BASE = os.environ.get("UPREP_BASE", "https://65.2.108.70.sslip.io")
CONTENT_SHEETS = [
    "Physics XI", "Physics XII", "Chemistry XI", "Chemistry XII",
    "Maths XI", "Maths XII", "Botany XI", "Botany XII", "Zoology XI", "Zoology XII",
]


def clean_topic(t: str) -> str:
    t = (t or "").strip()
    if t.lower().endswith(".mp4"):
        t = t[:-4].strip()
    return t


def parse_sheet(ws):
    """Return ordered list of (chapter_name, [(session, topic, url), ...])."""
    chapters = []
    index = {}
    current = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        vals = [("" if v is None else str(v)).strip() for v in row]
        if len(vals) < 4:
            vals += [""] * (4 - len(vals))
        chap, session, topic, url = vals[0], vals[1], vals[2], vals[3]
        if chap:
            if chap not in index:
                index[chap] = []
                chapters.append((chap, index[chap]))
            current = index[chap]
        if url and topic and current is not None:
            current.append((session, clean_topic(topic), url.strip()))
    # drop chapters with no videos
    return [(c, v) for c, v in chapters if v]


def summary():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    grand = 0
    for name in CONTENT_SHEETS:
        if name not in wb.sheetnames:
            continue
        chapters = parse_sheet(wb[name])
        vids = sum(len(v) for _, v in chapters)
        grand += vids
        print(f"{name:15s}  chapters={len(chapters):3d}  videos={vids:4d}")
    print(f"{'TOTAL':15s}  videos={grand}")


class Client:
    def __init__(self, base, user, pw):
        self.base = base
        self.s = requests.Session()
        r = self.s.post(f"{base}/api/auth/login",
                        json={"identifier": user, "password": pw}, timeout=30)
        r.raise_for_status()
        prof = r.json().get("result", {}).get("profile")
        if "uprep_auth" not in self.s.cookies.get_dict():
            raise SystemExit(f"Login failed (no cookie). Response: {r.text[:200]}")
        print(f"Logged in as {user} (profile={prof})")

    def folders_under(self, parent_id):
        params = {}
        if parent_id:
            params["parentId"] = parent_id
        r = self.s.get(f"{self.base}/api/cmds/content", params=params, timeout=30)
        r.raise_for_status()
        out = {}
        for res in r.json().get("resources", []):
            if res.get("type") == "FOLDER":
                out[res["title"].strip().lower()] = res["id"]
        return out

    def videos_in(self, folder_id):
        r = self.s.get(f"{self.base}/api/cmds/content",
                       params={"parentId": folder_id}, timeout=30)
        r.raise_for_status()
        return {res["title"].strip().lower()
                for res in r.json().get("resources", [])
                if res.get("type") == "VIDEO"}

    def get_or_create_folder(self, name, parent_id):
        existing = self.folders_under(parent_id)
        key = name.strip().lower()
        if key in existing:
            return existing[key], False
        r = self.s.post(f"{self.base}/api/cmds/content",
                        json={"kind": "folder", "name": name, "parentId": parent_id},
                        timeout=30)
        r.raise_for_status()
        return r.json()["id"], True

    def add_video(self, name, subject, url, folder_id):
        r = self.s.post(f"{self.base}/api/cmds/videos",
                        json={"name": name, "subject": subject, "url": url,
                              "folderId": folder_id}, timeout=30)
        if r.status_code != 200:
            return False, r.text[:160]
        return True, r.json().get("provider")

    def children(self, parent_id):
        params = {"parentId": parent_id} if parent_id else {}
        r = self.s.get(f"{self.base}/api/cmds/content", params=params, timeout=30)
        r.raise_for_status()
        return r.json().get("resources", [])

    def delete(self, res_id, res_type):
        self.s.delete(f"{self.base}/api/cmds/content",
                      params={"id": res_id, "type": res_type}, timeout=30)

    def wipe_folder(self, folder_id):
        """Recursively soft-delete a folder's contents, then the folder."""
        for r in self.children(folder_id):
            if r["type"] == "FOLDER":
                self.wipe_folder(r["id"])
            else:
                self.delete(r["id"], r["type"])
        self.delete(folder_id, "FOLDER")

    def wipe_subject(self, sheet):
        roots = self.folders_under(None)
        fid = roots.get(sheet.strip().lower())
        if fid:
            print(f"Reset: wiping existing '{sheet}' folder {fid} ...")
            self.wipe_folder(fid)


def import_sheet(c, wb, sheet, max_chapters=0, reset=False, delay=0.0):
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet}. Options: {CONTENT_SHEETS}")
    chapters = parse_sheet(wb[sheet])
    if max_chapters and max_chapters > 0:
        chapters = chapters[:max_chapters]

    if reset:
        c.wipe_subject(sheet)
    subject_folder, created = c.get_or_create_folder(sheet, None)
    print(f"\n########## {sheet} ##########")
    print(f"Subject folder '{sheet}': {subject_folder} ({'created' if created else 'exists'})",
          flush=True)

    total_added = total_skipped = total_failed = 0
    for chap, vids in chapters:
        cf, ccreated = c.get_or_create_folder(chap, subject_folder)
        print(f"\n  Chapter '{chap}' -> {cf} ({'created' if ccreated else 'exists'}), "
              f"{len(vids)} videos")
        # Group videos by session (preserving first-seen order); blank session
        # files straight into the chapter folder.
        session_folders = {}   # session label -> folder id
        session_have = {}      # folder id -> set of existing video titles
        for session, topic, url in vids:
            label = (session or "").strip()
            if label:
                if label not in session_folders:
                    sf, screated = c.get_or_create_folder(label, cf)
                    session_folders[label] = sf
                    session_have[sf] = c.videos_in(sf)
                    print(f"    · {label} -> {sf} ({'created' if screated else 'exists'})")
                target = session_folders[label]
            else:
                target = cf
                session_have.setdefault(cf, c.videos_in(cf))
            if topic.strip().lower() in session_have.get(target, set()):
                total_skipped += 1
                continue
            ok, info = c.add_video(topic, sheet, url, target)
            if ok:
                total_added += 1
                session_have.setdefault(target, set()).add(topic.strip().lower())
                if delay:
                    time.sleep(delay)
            else:
                total_failed += 1
                print(f"      ! FAILED {topic}: {info}", flush=True)
        print(f"    -> running: added={total_added} skipped={total_skipped} "
              f"failed={total_failed}", flush=True)
    print(f"DONE {sheet}: added={total_added} skipped={total_skipped} failed={total_failed}",
          flush=True)
    return total_added, total_skipped, total_failed


def do_import(sheet, max_chapters, user, pw, reset=False, delay=0.0):
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    c = Client(BASE, user, pw)
    import_sheet(c, wb, sheet, max_chapters, reset, delay)


def do_all(user, pw, delay, sheets):
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    c = Client(BASE, user, pw)
    g_add = g_skip = g_fail = 0
    for i, sheet in enumerate(sheets, 1):
        print(f"\n===== [{i}/{len(sheets)}] {sheet} =====", flush=True)
        a, s, f = import_sheet(c, wb, sheet, delay=delay)
        g_add += a; g_skip += s; g_fail += f
        print(f"===== cumulative: added={g_add} skipped={g_skip} failed={g_fail} =====",
              flush=True)
    print(f"\nALL DONE: added={g_add} skipped={g_skip} failed={g_fail}", flush=True)


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("summary")
    imp = sub.add_parser("import")
    imp.add_argument("--sheet", required=True)
    imp.add_argument("--max-chapters", type=int, default=0)
    imp.add_argument("--reset", action="store_true",
                     help="wipe the subject folder subtree before importing")
    imp.add_argument("--delay", type=float, default=0.0)
    alc = sub.add_parser("all")
    alc.add_argument("--delay", type=float, default=0.2)
    alc.add_argument("--sheets", nargs="*", default=CONTENT_SHEETS)
    args = ap.parse_args()

    if args.cmd == "summary":
        summary()
        return
    user = os.environ.get("UPREP_USER")
    pw = os.environ.get("UPREP_PASS")
    if not user or not pw:
        raise SystemExit("Set UPREP_USER and UPREP_PASS env vars.")
    if args.cmd == "import":
        do_import(args.sheet, args.max_chapters, user, pw, reset=args.reset, delay=args.delay)
    elif args.cmd == "all":
        do_all(user, pw, args.delay, args.sheets)


if __name__ == "__main__":
    main()
